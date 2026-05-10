import time
from pathlib import Path

from Bio import Entrez, SeqIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

DEFAULT_EMAIL = "your_email@example.com"


def prompt_required(prompt_text: str) -> str:
    while True:
        value = input(prompt_text).strip()
        if value:
            return value
        print("This field is required.")


def prompt_optional_int(prompt_text: str):
    value = input(prompt_text).strip()
    if not value:
        return None
    try:
        return int(value)
    except ValueError:
        print("Invalid number. It will be ignored.")
        return None


def build_query(gene_names, organism=None, taxid=None, length_min=None, length_max=None):
    gene_query = " OR ".join(f'({gene}[Gene Name])' for gene in gene_names)
    terms = [f"({gene_query})"]

    if organism:
        terms.append(f"({organism}[Organism])")
    if taxid:
        terms.append(f"(txid{taxid}[Organism:exp])")

    if length_min is not None or length_max is not None:
        min_len = 1 if length_min is None else max(length_min, 1)
        max_len = 1000000000 if length_max is None else max(length_max, min_len)
        terms.append(f"({min_len}:{max_len}[SLEN])")

    return " AND ".join(terms)


def fetch_records(database, query, max_sequences=None):
    search_handle = Entrez.esearch(db=database, term=query, retmax=0)
    search_record = Entrez.read(search_handle)
    search_handle.close()

    total_count = int(search_record.get("Count", 0))
    target_count = total_count if not max_sequences else min(total_count, max_sequences)

    if target_count == 0:
        return []

    search_handle = Entrez.esearch(db=database, term=query, retmax=target_count)
    search_record = Entrez.read(search_handle)
    search_handle.close()

    ids = search_record.get("IdList", [])
    if not ids:
        return []

    time.sleep(0.34)
    fetch_handle = Entrez.efetch(db=database, id=ids, rettype="fasta", retmode="text")
    records = list(SeqIO.parse(fetch_handle, "fasta"))
    fetch_handle.close()
    return records


def write_excel(records, output_xlsx: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Secuencias"

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="2E4057")
    centered = Alignment(horizontal="center", vertical="center")

    headers = ["Accession Number", "Name", "Length"]
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = centered

    even_fill = PatternFill("solid", start_color="EAF2FB")

    for row, record in enumerate(records, start=2):
        ws.cell(row=row, column=1, value=record.id)
        ws.cell(row=row, column=2, value=record.description)
        ws.cell(row=row, column=3, value=len(record.seq))

        if row % 2 == 0:
            for col in range(1, 4):
                ws.cell(row=row, column=col).fill = even_fill

        ws.cell(row=row, column=1).alignment = centered
        ws.cell(row=row, column=3).alignment = centered

        for col in range(1, 4):
            ws.cell(row=row, column=col).font = Font(name="Arial", size=10)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 90
    ws.column_dimensions["C"].width = 12
    ws.freeze_panes = "A2"

    wb.save(output_xlsx)


def main():
    print("=== NCBI Gene Downloader ===")
    print("Only gene name is required. Press Enter to skip optional fields.\n")

    raw_genes = prompt_required("Gene name(s) (comma-separated): ")
    gene_names = [gene.strip() for gene in raw_genes.split(",") if gene.strip()]

    database = input("NCBI database [nucleotide]: ").strip() or "nucleotide"
    organism = input("Organism name (optional): ").strip() or None
    taxid = input("TaxID (optional): ").strip() or None
    length_min = prompt_optional_int("Minimum sequence length (optional): ")
    length_max = prompt_optional_int("Maximum sequence length (optional): ")
    output_base = input("Output filename base [download]: ").strip() or "download"
    max_sequences = prompt_optional_int("Max sequences to download (optional, default unlimited): ")

    email = input("Email for NCBI Entrez (required by NCBI) [your_email@example.com]: ").strip() or DEFAULT_EMAIL
    Entrez.email = email

    if email == DEFAULT_EMAIL:
        print("WARNING: You are using a placeholder email. Replace it with your real email when possible.")

    output_base = output_base.removesuffix(".fasta").removesuffix(".xlsx")
    fasta_path = Path(f"{output_base}.fasta")
    excel_path = Path(f"{output_base}.xlsx")

    query = build_query(
        gene_names=gene_names,
        organism=organism,
        taxid=taxid,
        length_min=length_min,
        length_max=length_max,
    )

    print(f"\nSearching NCBI ({database})...")
    try:
        records = fetch_records(database=database, query=query, max_sequences=max_sequences)
    except Exception as exc:
        print(f"Error while contacting NCBI: {exc}")
        return

    SeqIO.write(records, fasta_path, "fasta")
    write_excel(records, excel_path)

    print(f"\nDone. {len(records)} sequences were downloaded.")
    print(f"FASTA: {fasta_path.resolve()}")
    print(f"Excel: {excel_path.resolve()}")


if __name__ == "__main__":
    main()
